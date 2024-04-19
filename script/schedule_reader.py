# schedule_reader.py

import openpyxl

def read_schedule(file_path, group_page):
     # Открываем файл Excel
    wb = openpyxl.load_workbook(file_path)
    
    # Получаем нужную страницу с расписанием для группы
    sheet = wb[group_page]
    
    # Словари для хранения расписания для каждой группы
    schedule_group1 = {}
    schedule_group2 = {}
    
    # Переменная для хранения текущего дня недели
    current_day = None
    
    # Дни недели, которые мы будем учитывать
    valid_days = ['Пн', 'Вт', 'Ср', 'Чт', 'Пт', 'Сб']
    
    # Проходимся по всем строкам на листе, начиная с 7 строки
    for row in range(7, sheet.max_row + 1):
        day = sheet.cell(row=row, column=1).value
        
        # Если встретился новый день недели
        if day and day in valid_days:
            current_day = day
            schedule_group1[current_day] = {}
            schedule_group2[current_day] = {}
        
        # Если текущего дня нет в списке дней недели, пропускаем обработку строки
        if not current_day:
            continue
        
        # Читаем данные пары
        pair = sheet.cell(row=row, column=2).value
        time_start_group1 = sheet.cell(row=row, column=3).value
        subject_group1 = sheet.cell(row=row, column=4).value
        room_group1 = sheet.cell(row=row, column=8).value
        
        time_start_group2 = sheet.cell(row=row, column=9).value
        subject_group2 = sheet.cell(row=row, column=10).value
        room_group2 = sheet.cell(row=row, column=14).value
        
        # Добавляем данные в расписание для текущей пары для каждой группы
        schedule_group1[current_day].setdefault(pair, {
            'time_start': time_start_group1,
            'subject': subject_group1,
            'room': room_group1
        })
        schedule_group2[current_day].setdefault(pair, {
            'time_start': time_start_group2,
            'subject': subject_group2,
            'room': room_group2
        })
    
    # Закрываем файл Excel
    wb.close()
    
    print("Schedule Group 1:", schedule_group1)
    print("Schedule Group 2:", schedule_group2)


    return schedule_group1, schedule_group2
